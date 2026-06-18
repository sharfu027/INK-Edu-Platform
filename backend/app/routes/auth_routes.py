# pyre-ignore-all-errors
"""
Authentication API routes.
Handles user registration, login, face verification, and profile access.
"""

import logging
from datetime import datetime, timezone, timedelta
from typing import List
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Form, Response
from fastapi.security import HTTPAuthorizationCredentials
from pydantic import BaseModel, Field
from bson import ObjectId

from app.config.database import get_database
from app.models.user import (
    UserRegisterRequest,
    UserLoginRequest,
    FaceVerifyRequest,
    RegisterResponse,
    AuthTokenResponse,
    FaceVerificationResponse,
    StandardResponse,
    CheckUserRequest,
    CompanySettings,
    EmployeeSettingsUpdate,
    SUPER_ADMIN_EMAIL,
    PERMANENT_ADMINS,
)

from typing import Optional, List, Any
from app.models.user import LocationData

class CompanySettings(BaseModel):
    hours_per_day: float
    hours_per_week: float
    hours_per_month: float
    hours_per_year: float
    weekly_off: str
    office_start_time: str
    grace_period_mins: int

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str = Field(..., min_length=8)

class AdminResetPasswordRequest(BaseModel):
    new_password: str = Field(..., min_length=8)

class KioskLogoutRequest(BaseModel):
    employee_id: str
    login_time: str
    duration_minutes: float
    face_image: Optional[str] = None
    challenge_frame: Optional[str] = None
    location: Optional[LocationData] = None

class LogoutRequest(BaseModel):
    location: Optional[LocationData] = None
from app.services.auth_service import AuthService
from app.middleware.auth_middleware import get_current_user, security
from app.models.user import ADMIN_PROFESSIONS


logger = logging.getLogger(__name__)

# Indian Standard Time (UTC+5:30) for attendance date/time alignment
IST = timezone(timedelta(hours=5, minutes=30))

router = APIRouter(prefix="/api/auth", tags=["Authentication"])


async def _is_admin(user_doc, db=None):
    """Check if user is admin. Super admin is always admin by email."""
    if not user_doc:
        return False
    
    # Super admin check by email (permanent, cannot be demoted)
    if user_doc.get("email", "").lower() in [e.lower() for e in PERMANENT_ADMINS]:
        # Auto-fix role if needed
        if user_doc.get("role") != "admin" and db:
            try:
                user_id = user_doc.get("_id") or user_doc.get("id")
                if user_id:
                    await db.users.update_one(
                        {"_id": ObjectId(str(user_id))},
                        {"$set": {"role": "admin"}}
                    )
            except Exception:
                pass
        return True
    
    # Direct role check (for users promoted by admin)
    if user_doc.get("role") == "admin":
        return True
    
    return False


# ──────────────────────────────────────────────
#  POST /api/auth/register
# ──────────────────────────────────────────────

@router.get("/settings", response_model=StandardResponse)
async def get_company_settings(db=Depends(get_database)):
    """Get the global company settings (without requiring authentication for registration display)."""
    settings_doc = await db.settings.find_one({"type": "company_config"})
    if settings_doc:
        settings_doc["_id"] = str(settings_doc["_id"])
        return StandardResponse(status=True, message="Settings fetched", data=settings_doc)
    return StandardResponse(
        status=True, 
        message="Defaults", 
        data={
            "hours_per_day": 8.0, 
            "hours_per_week": 40.0,
            "hours_per_month": 160.0,
            "hours_per_year": 1920.0,
            "weekly_off": "Sunday",
            "office_start_time": "09:00",
            "grace_period_mins": 30
        }
    )

@router.post("/settings", response_model=StandardResponse)
async def update_company_settings(
    settings: CompanySettings, 
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Update global company settings (Admin only). Also updates all existing employees to match new settings."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    from app.services.auth_service import AuthService
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only administrators can update company settings."
        )

    await db.settings.update_one(
        {"type": "company_config"},
        {
            "$set": {
                "hours_per_day": settings.hours_per_day,
                "hours_per_week": settings.hours_per_week,
                "hours_per_month": settings.hours_per_month,
                "hours_per_year": settings.hours_per_year,
                "weekly_off": settings.weekly_off,
                "office_start_time": settings.office_start_time,
                "grace_period_mins": settings.grace_period_mins,
                "updated_at": datetime.now(timezone.utc)
            }
        },
        upsert=True
    )
    
    # Cascade updates to all existing employees so they reflect the new company standard
    await db.users.update_many(
        {},
        {"$set": {
            "hours_per_day": settings.hours_per_day, 
            "hours_per_week": settings.hours_per_week,
            "hours_per_month": settings.hours_per_month,
            "hours_per_year": settings.hours_per_year,
            "weekly_off": settings.weekly_off,
            "office_start_time": settings.office_start_time,
            "grace_period_mins": settings.grace_period_mins
        }}
    )
    
    return StandardResponse(status=True, message="Settings updated and cascaded to all employees successfully")


# ──────────────────────────────────────────────
#  Password Reset (Forgot Password)
# ──────────────────────────────────────────────

@router.post("/change-password", response_model=StandardResponse)
async def change_user_password(
    request: ChangePasswordRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Securely change password for the authenticated user. Requires old password verification."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    
    auth_service = AuthService(db)
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    # Verify old password
    if not auth_service.verify_password(request.old_password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Incorrect old password")
        
    # Update with new password
    new_hash = auth_service.hash_password(request.new_password)
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return StandardResponse(status=True, message="Password updated successfully")


# ──────────────────────────────────────────────
#  Login History (per-user + admin search)
# ──────────────────────────────────────────────

@router.get("/login-history", response_model=StandardResponse)
async def get_login_history(
    search: str = "",
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """
    Return attendance punch history.
    - Regular users: only their own records.
    - Admin users: can search by employee name or ID to see any employee's records.
      If no search query, returns all employees' records.
    """
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")

    from app.services.auth_service import AuthService
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")

    is_admin = await _is_admin(user_doc, db)

    # 1. Fetch relevant users to get their details and login_sessions
    if is_admin:
        users_list = await db.users.find({}, {"employee_id": 1, "name": 1, "login_sessions": 1}).to_list(1000)
    else:
        users_list = [user_doc]

    emp_name_map = {u.get("employee_id"): u.get("name", "Unknown") for u in users_list if u.get("employee_id")}
    emp_ids = list(emp_name_map.keys())

    # 2. Fetch all attendance punch records for these users
    att_filter = {"employee_id": {"$in": emp_ids}}
    attendance_records = await db.attendance.find(att_filter).to_list(1000)

    history = []

    # 3. Process attendance records
    for r in attendance_records:
        eid = r.get("employee_id")
        if not eid:
            continue
        
        # Determine times
        pin = r.get("punch_in")
        pout = r.get("punch_out")
        
        # Calculate worked hours
        worked = r.get("worked_hours")
        if worked is None and pin and pout:
            try:
                pin_dt = datetime.fromisoformat(pin) if isinstance(pin, str) else pin
                pout_dt = datetime.fromisoformat(pout) if isinstance(pout, str) else pout
                if pin_dt.tzinfo is None:
                    pin_dt = pin_dt.replace(tzinfo=timezone.utc)
                if pout_dt.tzinfo is None:
                    pout_dt = pout_dt.replace(tzinfo=timezone.utc)
                worked = round((pout_dt - pin_dt).total_seconds() / 3600, 2)
            except Exception:
                worked = None

        # Geocoded login / logout location structure
        login_loc = r.get("login_location")
        logout_loc = r.get("logout_location")

        history.append({
            "employee_id": eid,
            "employee_name": emp_name_map.get(eid, "Unknown"),
            "date": r.get("date"),
            "punch_in": pin,
            "punch_out": pout,
            "login_location": login_loc,
            "logout_location": logout_loc,
            "status": "Completed" if pout else ("Active" if pin else "Incomplete"),
            "worked_hours": worked,
            "record_type": r.get("type", "attendance_punch"),
        })

    # 4. Process user login_sessions (Web sessions)
    for u in users_list:
        eid = u.get("employee_id")
        if not eid:
            continue
        name = u.get("name", "Unknown")
        sessions = u.get("login_sessions") or []
        for session in sessions:
            login_at = session.get("login_at")
            logout_at = session.get("logout_at")
            
            if isinstance(login_at, datetime):
                login_at = login_at.isoformat()
            if isinstance(logout_at, datetime):
                logout_at = logout_at.isoformat()
                
            date_str = ""
            if login_at:
                try:
                    date_str = login_at.split("T")[0]
                except Exception:
                    pass

            # Merge geocoded address fields so frontend formatLocation can read them!
            login_loc_merged = None
            if session.get("location"):
                login_loc_merged = {**session["location"]}
                if session.get("address"):
                    login_loc_merged.update(session["address"])
                    
            logout_loc_merged = None
            if session.get("logout_location"):
                logout_loc_merged = {**session["logout_location"]}
                if session.get("logout_address"):
                    logout_loc_merged.update(session["logout_address"])

            # Calculate worked hours
            worked = None
            if login_at and logout_at:
                try:
                    pin_dt = datetime.fromisoformat(login_at)
                    pout_dt = datetime.fromisoformat(logout_at)
                    if pin_dt.tzinfo is None:
                        pin_dt = pin_dt.replace(tzinfo=timezone.utc)
                    if pout_dt.tzinfo is None:
                        pout_dt = pout_dt.replace(tzinfo=timezone.utc)
                    worked = round((pout_dt - pin_dt).total_seconds() / 3600, 2)
                except Exception:
                    pass

            history.append({
                "employee_id": eid,
                "employee_name": name,
                "date": date_str,
                "punch_in": login_at,
                "punch_out": logout_at,
                "login_location": login_loc_merged,
                "logout_location": logout_loc_merged,
                "status": "Completed" if logout_at else "Active",
                "worked_hours": worked,
                "record_type": "web_session",
            })

    # Sort history by punch_in descending
    history.sort(key=lambda x: x.get("punch_in") or "", reverse=True)

    # 5. Filter history list based on search query if applicable
    if search.strip():
        q = search.strip().lower()
        history = [
            h for h in history
            if q in (h.get("employee_name") or "").lower()
            or q in (h.get("employee_id") or "").lower()
            or q in (h.get("date") or "").lower()
            or q in (h.get("status") or "").lower()
            or q in (h.get("record_type") or "").lower()
        ]

    # Deduplicate entries that might be identical (same employee, same punch_in down to the minute)
    seen = set()
    deduped_history = []
    for h in history:
        pin_time = h.get("punch_in") or ""
        if pin_time:
            minute_key = pin_time[:16] # e.g. "2026-05-17T13:29"
            key = (h.get("employee_id"), minute_key)
            if key in seen:
                continue
            seen.add(key)
        deduped_history.append(h)

    return StandardResponse(status=True, message=f"Found {len(deduped_history)} records", data=deduped_history)


# ──────────────────────────────────────────────
#  Admin Employee Management
# ──────────────────────────────────────────────

@router.get("/admin/employees", response_model=StandardResponse)
async def admin_list_employees(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """List all employees with their face/location/role settings. Admin only."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    employees = []
    async for u in db.users.find({}, {"password_hash": 0, "face_embeddings": 0}):
        u["_id"] = str(u["_id"])
        is_super = u.get("email", "").lower() in [e.lower() for e in PERMANENT_ADMINS]
        employees.append({
            "_id": u["_id"],
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "phone": u.get("phone", ""),
            "employee_id": u.get("employee_id", ""),
            "designation": u.get("designation", ""),
            "profession": u.get("profession", ""),
            "joining_date": u.get("joining_date", ""),
            "role": u.get("role", "user"),
            "skip_face": u.get("skip_face", False),
            "skip_location": u.get("skip_location", False),
            "is_enabled": u.get("is_enabled", False),
            "is_super_admin": is_super,
            "created_at": str(u.get("created_at", "")),
            "liveness_verified": u.get("liveness_verified", False),
            "hours_per_day": u.get("hours_per_day", 8),
            "weekly_off": u.get("weekly_off", "Sunday"),
            "insurance_id": u.get("insurance_id", ""),
            "insurance_provider": u.get("insurance_provider", ""),
            "plain_password": u.get("password_plain") or u.get("plain_password", ""),
            "govt_id_filename": u.get("govt_id_card", {}).get("filename", "") if u.get("govt_id_card") else "",
            "insurance_filename": u.get("insurance_card", {}).get("filename", "") if u.get("insurance_card") else "",
        })

    return StandardResponse(status=True, message=f"{len(employees)} employees found", data=employees)


@router.put("/admin/employee/{employee_id}/settings", response_model=StandardResponse)
async def admin_update_employee_settings(
    employee_id: str,
    settings: EmployeeSettingsUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Update an employee's skip_face, skip_location, role, and insurance details. Admin only."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    from bson import ObjectId
    if ObjectId.is_valid(employee_id):
        target = await db.users.find_one({"_id": ObjectId(employee_id)})
    else:
        target = await db.users.find_one({"employee_id": employee_id})
        
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    if target.get("email", "").lower() in [e.lower() for e in PERMANENT_ADMINS]:
        if settings.role and settings.role != "admin":
            raise HTTPException(status_code=400, detail="Cannot demote the super admin")

    update_fields: dict[str, Any] = {"updated_at": datetime.now(timezone.utc)}
    if settings.skip_face is not None:
        update_fields["skip_face"] = settings.skip_face
    if settings.skip_location is not None:
        update_fields["skip_location"] = settings.skip_location
    if settings.is_enabled is not None:
        update_fields["is_enabled"] = settings.is_enabled
    if settings.role is not None:
        update_fields["role"] = settings.role
    if settings.insurance_id is not None:
        update_fields["insurance_id"] = settings.insurance_id
    if settings.insurance_provider is not None:
        update_fields["insurance_provider"] = settings.insurance_provider

    await db.users.update_one({"_id": target["_id"]}, {"$set": update_fields})
    logger.info(f"Admin updated employee {target.get('employee_id', employee_id)}: {update_fields}")
    return StandardResponse(status=True, message=f"Employee {target.get('name', employee_id)} settings updated successfully")


@router.delete("/admin/employee/{employee_id}", response_model=StandardResponse)
async def admin_delete_employee(
    employee_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Hard delete an employee record. Admin only."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    from bson import ObjectId
    if ObjectId.is_valid(employee_id):
        target = await db.users.find_one({"_id": ObjectId(employee_id)})
    else:
        target = await db.users.find_one({"employee_id": employee_id})
        
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Prevent deleting the super admin
    if target.get("email", "").lower() in [e.lower() for e in PERMANENT_ADMINS]:
        raise HTTPException(status_code=403, detail="Cannot delete the Super Admin account")

    await db.users.delete_one({"_id": target["_id"]})
    # Also delete their attendance records
    if target.get("employee_id"):
        await db.attendance.delete_many({"employee_id": target["employee_id"]})
        
    logger.info(f"Admin deleted employee {target.get('employee_id', employee_id)}")
    return StandardResponse(status=True, message=f"Employee {target.get('name', employee_id)} deleted permanently")
    
@router.post("/admin/employee/{employee_id}/reset-password", response_model=StandardResponse)
async def admin_reset_employee_password(
    employee_id: str,
    request: AdminResetPasswordRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Reset an employee's password. Admin only."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    from bson import ObjectId
    if ObjectId.is_valid(employee_id):
        target = await db.users.find_one({"_id": ObjectId(employee_id)})
    else:
        target = await db.users.find_one({"employee_id": employee_id})
        
    if not target:
        return StandardResponse(status=False, message="Failed to reset password. Employee might not exist.")

    success = await auth_service.admin_reset_password(str(target["_id"]), request.new_password)
    if success:
        return StandardResponse(status=True, message=f"Password for employee {target.get('name', employee_id)} has been reset successfully.")
    else:
        return StandardResponse(status=False, message="Failed to reset password.")

def calc_late_early(login_str: str, office_start_str: str = "09:00", grace_mins: int = 30):
    if not login_str or login_str == "-":
        return "", "", 0, 0
    try:
        t_login = datetime.strptime(login_str, "%I:%M %p")
        login_mins = t_login.hour * 60 + t_login.minute
        
        t_office = datetime.strptime(office_start_str, "%H:%M")
        office_start_mins = t_office.hour * 60 + t_office.minute
        
        late_mins_calc = 0
        early_mins_calc = 0
        
        if login_mins > (office_start_mins + grace_mins):
            late_mins_calc = login_mins - office_start_mins
        elif login_mins < office_start_mins:
            early_mins_calc = office_start_mins - login_mins
            
        def format_mins(m):
            if m == 0: return ""
            h = m // 60
            mins = m % 60
            if h > 0: return f"{h}h {mins}m"
            return f"{mins}m"
            
        return format_mins(late_mins_calc), format_mins(early_mins_calc), late_mins_calc, early_mins_calc
    except Exception:
        return "", "", 0, 0

@router.get("/attendance/export")
async def export_attendance_excel(
    report_type: str = "consolidated",
    month: Optional[int] = None,
    year: Optional[int] = None,
    department: Optional[str] = None,
    employee_id: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Export attendance report as Excel file. Admin only."""
    from fastapi.responses import StreamingResponse
    import io, calendar

    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    now = datetime.now(timezone.utc)
    target_month = month or now.month
    target_year = year or now.year
    days_in_month = calendar.monthrange(target_year, target_month)[1]

    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws: Any = wb.active
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    user_filter = {}
    if employee_id:
        user_filter["employee_id"] = employee_id
    elif department:
        user_filter["$or"] = [
            {"profession": {"$regex": department, "$options": "i"}},
            {"designation": {"$regex": department, "$options": "i"}},
        ]
    all_users = await db.users.find(user_filter, {"password_hash": 0, "face_embeddings": 0}).to_list(length=None)
    settings_doc = await db.settings.find_one({"type": "company_config"}) or {}
    global_weekly_off = settings_doc.get("weekly_off", "Sunday")
    office_start_time = settings_doc.get("office_start_time", "09:00")
    grace_period_mins = settings_doc.get("grace_period_mins", 30)

    date_prefix = f"{target_year}-{str(target_month).zfill(2)}"

    # OPTIMIZATION: Fetch all punches for the month at once
    all_punches = await db.attendance.find({
        "type": {"$in": ["punch_in", "punch", None]},
        "date": {"$regex": f"^{date_prefix}"}
    }).to_list(10000)

    punch_map = {}
    for p in all_punches:
        key = (p.get("employee_id"), p.get("date"))
        punch_map[key] = p

    if report_type == "details":
        # New Grid structure matching the third image!
        ws.title = "All Details"
        
        # Row 1 Headers
        c_date_lbl = ws.cell(row=1, column=2, value="Date")
        c_date_lbl.font = header_font
        c_date_lbl.fill = header_fill
        c_date_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_date_lbl.border = thin_border
        
        for day in range(1, days_in_month + 1):
            c_day = ws.cell(row=1, column=day + 2, value=day)
            c_day.font = header_font
            c_day.fill = header_fill
            c_day.alignment = Alignment(horizontal="center", vertical="center")
            c_day.border = thin_border
            
        # Row 2 Headers
        c_name_lbl = ws.cell(row=2, column=1, value="name")
        c_name_lbl.font = header_font
        c_name_lbl.fill = header_fill
        c_name_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_name_lbl.border = thin_border
        
        # Border empty header cells A1 and B2
        for r, col in [(1, 1), (2, 2)]:
            c_empty = ws.cell(row=r, column=col, value="")
            c_empty.border = thin_border
            
        # Border Row 2 day columns to keep it neat
        for day in range(1, days_in_month + 1):
            c_empty = ws.cell(row=2, column=day + 2, value="")
            c_empty.border = thin_border
            
        # Row 3 onwards: data
        row = 3
        for u in all_users:
            eid = u.get("employee_id", "")
            
            # Column 1 (A): Employee Name
            c_emp_name = ws.cell(row=row, column=1, value=u.get("name", ""))
            c_emp_name.font = Font(bold=True, size=10)
            c_emp_name.alignment = Alignment(horizontal="left", vertical="center")
            c_emp_name.border = thin_border
            
            # Column 2 (B): Empty
            c_b = ws.cell(row=row, column=2, value="")
            c_b.border = thin_border
            
            # Column 3 onwards: day statuses
            for day in range(1, days_in_month + 1):
                date_str = f"{date_prefix}-{str(day).zfill(2)}"
                punch = punch_map.get((eid, date_str))
                status = "P" if punch else "A"
                c_status = ws.cell(row=row, column=day + 2, value=status)
                c_status.alignment = Alignment(horizontal="center", vertical="center")
                c_status.border = thin_border
                
                # Apply E2EFDA (light green) for P, FCE4D6 (light red) for A
                if status == "P":
                    c_status.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    c_status.font = Font(color="375623", bold=True)
                else:
                    c_status.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    c_status.font = Font(color="C65911", bold=True)
            row += 1
            
    else:
        # Consolidated Report (original logic)
        ws.title = "Consolidated Report"
        headers = ["Employee", "Employee ID", "Department", "Designation", "Days Present", "Days Absent", "Total Hours", "Target Hours/Day", "OT Hours", "Less Hours", "Total Late", "Total Early"]
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        row = 2
        for u in all_users:
            eid = u.get("employee_id", "")
            emp_hours = u.get("hours_per_day", 8)
            emp_weekly_off = u.get("weekly_off") or global_weekly_off
            
            present, total_h, working_days = 0, 0.0, 0
            total_late_mins, total_early_mins = 0, 0
            for day in range(1, days_in_month + 1):
                date_str = f"{date_prefix}-{str(day).zfill(2)}"
                day_name = datetime(target_year, target_month, day).strftime("%A")
                if day_name.lower() != emp_weekly_off.lower():
                    working_days += 1
                punch = punch_map.get((eid, date_str))
                if punch:
                    present += 1
                    login_t = punch.get("punch_in", "")
                    _, _, l_mins, e_mins = calc_late_early(login_t, office_start_time, grace_period_mins)
                    total_late_mins += l_mins
                    total_early_mins += e_mins
                    
                    dur = punch.get("duration", "")
                    if dur and "h" in str(dur):
                        try:
                            p = str(dur).replace("(active)","").strip().split("h")
                            total_h += int(p[0].strip()) + (int(p[1].replace("m","").strip())/60 if len(p)>1 and p[1].strip().replace("m","").strip() else 0)
                        except Exception:
                            pass
            absent = max(working_days - present, 0)
            target_h_monthly = round(working_days * emp_hours, 1)
            ot = round(total_h - target_h_monthly, 1) if total_h > target_h_monthly else 0
            deficit = round(target_h_monthly - total_h, 1) if total_h < target_h_monthly else 0
            
            def format_m(m):
                if m == 0: return ""
                h = m // 60
                mins = m % 60
                if h > 0: return f"{h}h {mins}m"
                return f"{mins}m"
                
            vals = [u.get("name",""), eid, u.get("profession","Employee"), u.get("designation",""), present, absent, round(total_h,1), emp_hours, ot or "", deficit or "", format_m(total_late_mins), format_m(total_early_mins)]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = thin_border
            row += 1

    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"attendance_{report_type}_{target_year}_{str(target_month).zfill(2)}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/check-user", response_model=StandardResponse)
async def check_user(request: CheckUserRequest, db=Depends(get_database)):
    """Check if email or phone is already registered."""
    email_exists = await db.users.find_one({"email": request.email})
    phone_exists = await db.users.find_one({"phone": request.phone})

    if email_exists and phone_exists:
        return StandardResponse(status=False, message="Email and phone number are already registered")
    elif email_exists:
        return StandardResponse(status=False, message="Email already registered")
    elif phone_exists:
        return StandardResponse(status=False, message="Phone number already registered")

    return StandardResponse(status=True, message="User not registered")



@router.post("/register", response_model=RegisterResponse, status_code=status.HTTP_201_CREATED)
async def register_user(request: UserRegisterRequest, db=Depends(get_database)):
    """
    Register a new user with face recognition.

    Requires:
    - Full name, email, phone, password
    - 4 face images (front, left, right, up/down) as base64 strings
    - Liveness detection is performed automatically
    """
    auth_service = AuthService(db)

    success, message, user_id = await auth_service.register_user(
        name=request.name,
        email=request.email,
        phone=request.phone,
        password=request.password,
        designation=request.designation,
        profession=request.profession,
        joining_date=request.joining_date,
        face_images=request.face_images,
        location=request.location.model_dump() if request.location else None,
    )

    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=message,
        )

    # Fetch the auto-generated employee_id for QR code generation
    new_user = await db.users.find_one({"_id": ObjectId(user_id)})
    emp_id = new_user.get("employee_id", "") if new_user else ""
    
    logger.info(f"New user registered: {user_id} (employee_id={emp_id})")
    return RegisterResponse(status=True, message=message, user_id=user_id, employee_id=emp_id)


# ──────────────────────────────────────────────
#  POST /api/auth/login
# ──────────────────────────────────────────────

@router.post("/login", response_model=AuthTokenResponse)
async def login_user(request: UserLoginRequest, db=Depends(get_database)):
    """
    Login with email and password.

    Returns JWT tokens but face verification is still required.
    Client must call /api/auth/verify-face next.
    """
    auth_service = AuthService(db)

    success, message, token_data = await auth_service.login_with_password(
        email=request.email,
        password=request.password,
        location=request.location.model_dump() if request.location else None,
    )

    if not success or not token_data:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=message,
        )

    logger.info(f"User logged in (password verified): {token_data['user_id']}")

    # Fetch user doc to check skip_face flag
    user_doc = await db.users.find_one({"_id": ObjectId(token_data["user_id"])})
    skip_face = user_doc.get("skip_face", False) if user_doc else False
    skip_location = user_doc.get("skip_location", False) if user_doc else False

    return AuthTokenResponse(
        access_token=token_data["access_token"],
        refresh_token=token_data["refresh_token"],
        token_type=token_data.get("token_type", "bearer"),
        requires_face_verification=not skip_face,
        user_id=token_data["user_id"],
    )


# ──────────────────────────────────────────────
#  POST /api/auth/verify-face
# ──────────────────────────────────────────────

@router.post("/verify-face", response_model=FaceVerificationResponse)
async def verify_face(
    request: FaceVerifyRequest,
    db=Depends(get_database),
):
    """
    Verify a user's face against stored embeddings.

    Called after password login or as standalone face login.
    Captures a live face image and compares it with stored embeddings.
    Enforces GPS location check against registered location.
    """
    from app.services.auth_service import haversine_distance, LOCATION_RADIUS_M
    from bson import ObjectId

    auth_service = AuthService(db)

    # ── Location check for verify-face ──
    user_doc = await db.users.find_one({"_id": ObjectId(request.user_id)})
    if user_doc:
        skip_location = user_doc.get("skip_location", False)
        reg_loc = user_doc.get("registered_location")
        login_loc = request.location.model_dump() if request.location else None

        if not skip_location and reg_loc and login_loc:
            dist = haversine_distance(
                reg_loc["latitude"], reg_loc["longitude"],
                login_loc["latitude"], login_loc["longitude"],
            )
            if dist > LOCATION_RADIUS_M:
                from app.utils.geocoding import reverse_geocode
                reg_addr = await reverse_geocode(reg_loc["latitude"], reg_loc["longitude"])
                curr_addr = await reverse_geocode(login_loc["latitude"], login_loc["longitude"])

                def format_addr(addr_doc, lat, lng):
                    if not addr_doc or "fallback" in addr_doc:
                        return f"{lat:.6f}, {lng:.6f}"
                    parts = [
                        addr_doc.get("road"),
                        addr_doc.get("suburb") or addr_doc.get("area"),
                        addr_doc.get("city"),
                        addr_doc.get("state"),
                        addr_doc.get("country"),
                        addr_doc.get("pincode")
                    ]
                    valid_parts = [p for p in parts if p is not None and str(p).strip()]
                    return ", ".join(valid_parts) if valid_parts else addr_doc.get("display_name", f"{lat:.6f}, {lng:.6f}")

                reg_display = f"({reg_loc['latitude']:.6f}, {reg_loc['longitude']:.6f}) - {format_addr(reg_addr, reg_loc['latitude'], reg_loc['longitude'])}"
                curr_display = f"({login_loc['latitude']:.6f}, {login_loc['longitude']:.6f}) - {format_addr(curr_addr, login_loc['latitude'], login_loc['longitude'])}"

                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=(
                        f"LOGIN FAILED — Location mismatch! "
                        f"You are {dist:.0f}m away from your registered location. "
                        f"Max allowed: {LOCATION_RADIUS_M}m. "
                        f"Registered: {reg_display}. "
                        f"Current: {curr_display}. "
                        f"You can only login from your registered location. "
                        f"To login from this new location, you must register a new account first."
                    ),
                )
        elif not skip_location and reg_loc and not login_loc:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Location is required for login. Please enable GPS/location services.",
            )

    is_verified, message, confidence = await auth_service.verify_face(
        user_id=request.user_id,
        face_image=request.face_image,
        challenge_frame=request.challenge_frame,
    )

    if is_verified:
        logger.info(f"Face verified for user: {request.user_id}")
    else:
        logger.warning(f"Face verification failed for user: {request.user_id}")

    return FaceVerificationResponse(
        status=is_verified,
        message=message,
        confidence=confidence,
    )


# ──────────────────────────────────────────────
#  POST /api/auth/face-login
# ──────────────────────────────────────────────

@router.post("/face-login")
async def face_login(request: FaceVerifyRequest, db=Depends(get_database)):
    """
    Login using face recognition only.
    Searches all users and matches the face embedding.
    Enforces location check against registered location.
    """
    from app.services.face_recognition import face_service
    from app.utils.encryption import decrypt_embeddings
    from app.services.auth_service import haversine_distance, LOCATION_RADIUS_M
    from bson import ObjectId

    # Extract embedding from the provided face image (strict quality checks + full-face validation)
    live_embedding, reason = face_service.extract_embedding_with_reason(request.face_image, strict=True)
    if live_embedding is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=reason,
        )

    # If user_id is provided, verify against that specific user
    if request.user_id:
        auth_service = AuthService(db)

        # Resolve the identifier (could be email or ObjectId)
        resolved_id = await auth_service.resolve_user_id(request.user_id)
        if not resolved_id:
            raise HTTPException(status_code=404, detail="User not found")

        # ── Location check for face-login ──
        user_doc = await db.users.find_one({"_id": ObjectId(resolved_id)})
        if user_doc:
            reg_loc = user_doc.get("registered_location")
            login_loc = request.location.model_dump() if request.location else None

            if reg_loc and login_loc:
                dist = haversine_distance(
                    reg_loc["latitude"], reg_loc["longitude"],
                    login_loc["latitude"], login_loc["longitude"],
                )
                if dist > LOCATION_RADIUS_M:
                    from app.utils.geocoding import reverse_geocode
                    reg_addr = await reverse_geocode(reg_loc["latitude"], reg_loc["longitude"])
                    curr_addr = await reverse_geocode(login_loc["latitude"], login_loc["longitude"])
                    
                    def format_addr(addr_doc, lat, lng):
                        if not addr_doc or "fallback" in addr_doc:
                            return f"{lat:.6f}, {lng:.6f}"
                        parts = [
                            addr_doc.get("road"),
                            addr_doc.get("suburb") or addr_doc.get("area"),
                            addr_doc.get("city"),
                            addr_doc.get("state"),
                            addr_doc.get("country"),
                            addr_doc.get("pincode")
                        ]
                        valid_parts = [p for p in parts if p is not None and str(p).strip()]
                        return ", ".join(valid_parts) if valid_parts else addr_doc.get("display_name", f"{lat:.6f}, {lng:.6f}")

                    email_reg_str = format_addr(reg_addr, reg_loc['latitude'], reg_loc['longitude'])
                    email_curr_str = format_addr(curr_addr, login_loc['latitude'], login_loc['longitude'])
                    
                    reg_display = f"({reg_loc['latitude']:.6f}, {reg_loc['longitude']:.6f}) - {reg_addr.get('display_name', 'Location')}"
                    curr_display = f"({login_loc['latitude']:.6f}, {login_loc['longitude']:.6f}) - {curr_addr.get('display_name', 'Location')}"
                    

                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail=(
                            f"LOGIN FAILED — Location mismatch! "
                            f"You are {dist:.0f}m away from your registered location. "
                            f"Max allowed: {LOCATION_RADIUS_M}m. "
                            f"Registered: {reg_display}. "
                            f"Current: {curr_display}. "
                            f"You can only login from your registered location. "
                            f"To login from this new location, you must register a new account first."
                        ),
                    )
            elif reg_loc and not login_loc:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Location is required for login. Please enable GPS/location services.",
                )
            elif not reg_loc and login_loc:
                logger.info(f"Legacy face-login for {resolved_id}. Setting current location as registered.")
                await db.users.update_one(
                    {"_id": ObjectId(resolved_id)},
                    {"$set": {"registered_location": login_loc}}
                )
            elif not reg_loc and not login_loc:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Location is required to secure your legacy account. Please enable GPS.",
                )

        is_verified, message, confidence = await auth_service.verify_face(
            user_id=resolved_id,
            face_image=request.face_image,
            challenge_frame=request.challenge_frame,
        )

        if not is_verified:

            return FaceVerificationResponse(status=False, message=message, confidence=confidence)

        user = await auth_service.get_user_by_id(resolved_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        access_token = AuthService.create_access_token(resolved_id, user["email"])
        refresh_token = AuthService.create_refresh_token(resolved_id)

        # Record login session for face-login users
        login_loc = request.location.model_dump() if request.location else None
        await auth_service._record_login(resolved_id, login_loc)
        


        return {
            "status": True,
            "message": "Face Verified",
            "confidence": confidence,
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer",
            "user_id": resolved_id,
        }

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="User ID is required for face login",
    )


# ──────────────────────────────────────────────
#  GET /api/auth/profile
# ──────────────────────────────────────────────

@router.get("/profile", response_model=StandardResponse)
async def get_profile(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database),
):
    """
    Get the authenticated user's profile.
    Requires a valid JWT access token.
    """
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    auth_service = AuthService(db)
    user = await auth_service.get_user_by_id(user_id)

    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found",
        )

    # Auto-fix role for users whose profession qualifies as admin
    # but whose role was set to "user" before the admin fix was deployed
    if user.get("role") != "admin":
        is_admin = await _is_admin(user, db)
        if is_admin:
            user["role"] = "admin"
            
    # ONE-TIME FIX for srujay@gmail.com if they registered before the auto-admin update
    if user.get("email", "").lower() == "srujay@gmail.com" and user.get("role") != "admin":
        if not user.get("srujay_admin_fixed"):
            try:
                await db.users.update_one(
                    {"_id": ObjectId(user_id)},
                    {"$set": {"role": "admin", "srujay_admin_fixed": True}}
                )
                user["role"] = "admin"
                logger.info(f"Auto-promoted srujay@gmail.com to admin (one-time fix).")
            except Exception as e:
                logger.error(f"Failed to auto-promote srujay: {e}")

    # MIGRATION FOR LEGACY ACCOUNTS: Auto-generate employee_id if missing
    if not user.get("employee_id"):
        import uuid
        new_employee_id = f"EMP-{uuid.uuid4().hex[:6].upper()}"
        logger.info(f"Legacy account {user_id} has no employee_id. Auto-generated {new_employee_id}.")
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"employee_id": new_employee_id}}
        )
        user["employee_id"] = new_employee_id

    return StandardResponse(status=True, message="Profile retrieved", data=user)


# ──────────────────────────────────────────────
#  PUT /api/auth/profile-photo
# ──────────────────────────────────────────────

class UpdateProfilePhotoRequest(BaseModel):
    """Schema for updating profile photo."""
    profile_photo: str = Field(..., description="Base64-encoded profile photo string")


@router.put("/profile-photo", response_model=StandardResponse)
async def update_profile_photo(
    request: UpdateProfilePhotoRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database),
):
    """
    Update the profile photo for the authenticated user.
    """
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"profile_photo": request.profile_photo, "updated_at": datetime.now(timezone.utc)}}
    )
    return StandardResponse(status=True, message="Profile photo updated successfully")


# ──────────────────────────────────────────────
#  PUT /api/auth/update-face
# ──────────────────────────────────────────────

class UpdateFaceRequest(BaseModel):
    """Schema for updating face data."""
    face_images: List[str] = Field(
        ...,
        min_length=4, max_length=4,
        description="4 base64-encoded face images: [front, left, right, up/down]"
    )


@router.put("/update-face", response_model=StandardResponse)
async def update_face_data(
    request: UpdateFaceRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database),
):
    """
    Add or update face data for an authenticated user.
    Useful when user registered without face data or wants to re-enroll.
    Requires a valid JWT access token.
    """
    from app.services.face_recognition import face_service
    from app.utils.encryption import encrypt_embeddings
    from datetime import datetime

    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    # Perform liveness check
    logger.info(f"Performing liveness check for face update (user: {user_id})")
    is_live, liveness_msg = face_service.perform_liveness_check(request.face_images)
    if not is_live:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Liveness verification failed: {liveness_msg}",
        )

    # Extract embeddings
    logger.info("Extracting face embeddings for update...")
    embeddings, errors = face_service.extract_multiple_embeddings(request.face_images)

    if errors:
        logger.warning(f"Some face images failed during update: {'; '.join(errors)}")

    if len(embeddings) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not extract enough face embeddings. Ensure your face is well-lit and clearly visible.",
        )

    # Encrypt and store
    encrypted_embeddings = encrypt_embeddings(embeddings)

    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {
            "$set": {
                "face_embeddings": encrypted_embeddings,
                "liveness_verified": True,
                "updated_at": datetime.now(timezone.utc),
            }
        },
    )

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    logger.info(f"Face data updated for user: {user_id} ({len(embeddings)} embeddings)")
    return StandardResponse(
        status=True,
        message=f"Face data updated successfully ({len(embeddings)} views enrolled)",
    )


# ──────────────────────────────────────────────
#  GET /api/auth/health
# ──────────────────────────────────────────────

@router.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "Face Auth API"}


# ──────────────────────────────────────────────
#  Logout / Auth Status (Protected)
# ──────────────────────────────────────────────

@router.post("/logout", response_model=StandardResponse)
async def logout_user(
    request: Optional[LogoutRequest] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database),
):
    """Record logout time and location for the authenticated user."""
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    auth_service = AuthService(db)
    loc_dict = request.location.model_dump() if request and request.location else None
    await auth_service.record_logout(user_id, loc_dict)

    logger.info(f"User logged out: {user_id}")
    return StandardResponse(status=True, message="Logged out successfully")


# ──────────────────────────────────────────────
#  Logout Kiosk API (Public/Tablet usage)
# ──────────────────────────────────────────────

@router.get("/kiosk/{employee_id}", response_model=StandardResponse)
async def kiosk_get_employee(employee_id: str, db=Depends(get_database)):
    """Fetch employee info publicly for the Logout Kiosk."""
    user = await db.users.find_one({"employee_id": employee_id})
    if not user:
        # Prevent leaking email/phone if user tries searching, only match exact employee ID
        return StandardResponse(status=False, message="Employee ID not found")
        
    return StandardResponse(status=True, message="Found", data={
        "name": user.get("name"),
        "designation": user.get("designation"),
        "employee_id": user.get("employee_id"),
        "hours_per_day": user.get("hours_per_day", 8),
        "hours_per_week": user.get("hours_per_week", 40),
        "hours_per_month": user.get("hours_per_month", 160),
        "hours_per_year": user.get("hours_per_year", 1920),
        "last_login_at": user.get("last_login_at"),
        "skip_face": user.get("skip_face", False),
        "skip_location": user.get("skip_location", False),
    })

@router.post("/kiosk/logout", response_model=StandardResponse)
async def kiosk_logout_employee(request: KioskLogoutRequest, db=Depends(get_database)):
    """Process logout from Kiosk. Also closes the attendance punch record."""
    user = await db.users.find_one({"employee_id": request.employee_id})
    if not user:
        return StandardResponse(status=False, message="Employee ID not found")

    # Enforce Face + Location match
    await _verify_punch_face_location(user, request.face_image, request.challenge_frame, request.location, db)

    now_utc = datetime.now(timezone.utc)
    now_ist = now_utc.astimezone(IST)
    today_str = now_ist.strftime("%Y-%m-%d")
    target_hours = user.get("hours_per_day", 8)

    # Find today's open punch-in record
    open_punch = await db.attendance.find_one({
        "employee_id": request.employee_id,
        "date": today_str,
        "type": "attendance_punch",
        "punch_out": None
    })

    # Check working hours (warning only, don't block logout)
    hours_warning = None
    if open_punch and open_punch.get("punch_in"):
        try:
            pin_str = open_punch["punch_in"]
            pin_dt = datetime.fromisoformat(pin_str) if isinstance(pin_str, str) else pin_str
            if pin_dt.tzinfo is None:
                pin_dt = pin_dt.replace(tzinfo=IST)
            worked_hours = (now_ist - pin_dt).total_seconds() / 3600
            if worked_hours < target_hours:
                remaining = target_hours - worked_hours
                rh = int(remaining)
                rm = int((remaining - rh) * 60)
                hours_warning = f"Warning: You worked {int(worked_hours)}h {int((worked_hours % 1) * 60)}m — {rh}h {rm}m less than required {target_hours}h."
            elif worked_hours > target_hours:
                extra = worked_hours - target_hours
                eh = int(extra)
                em = int((extra - eh) * 60)
                hours_warning = f"Great job! You worked {eh}h {em}m extra beyond the required {target_hours}h."
        except Exception as e:
            logger.warning(f"Could not check working hours: {e}")

    # Close the attendance punch record
    duration_str = None
    if open_punch:
        try:
            pin_str = open_punch["punch_in"]
            pin_dt = datetime.fromisoformat(pin_str) if isinstance(pin_str, str) else pin_str
            if pin_dt.tzinfo is None:
                pin_dt = pin_dt.replace(tzinfo=IST)
            diff = now_ist - pin_dt
            hours = int(diff.total_seconds() // 3600)
            minutes = int((diff.total_seconds() % 3600) // 60)
            duration_str = f"{hours}h {minutes}m"
        except Exception:
            duration_str = "N/A"

        # Geocode for address
        loc_dict = request.location.model_dump() if request.location else None
        if loc_dict:
            from app.utils.geocoding import reverse_geocode
            addr_res = await reverse_geocode(loc_dict["latitude"], loc_dict["longitude"])
            if addr_res:
                loc_dict["address"] = addr_res.get("display_name") or addr_res.get("address") or loc_dict.get("address")
                for k, v in addr_res.items():
                    if k not in loc_dict:
                        loc_dict[k] = v

        await db.attendance.update_one(
            {"_id": open_punch["_id"]},
            {"$set": {
                "punch_out": now_ist.isoformat(),
                "duration": duration_str,
                "status": "Logged Out",
                "logout_location": loc_dict
            }}
        )

    # Also create a kiosk_checkout record for backward compatibility
    attendance_record = {
        "employee_id": request.employee_id,
        "name": user.get("name"),
        "login_time_entered": request.login_time,
        "logout_time": now_utc,
        "duration_minutes": request.duration_minutes,
        "target_hours": target_hours,
        "type": "kiosk_checkout",
        "created_at": now_utc
    }
    await db.attendance.insert_one(attendance_record)

    await db.users.update_one(
        {"employee_id": request.employee_id},
        {"$set": {"last_logout_at": now_utc}}
    )
    duration_str_final = duration_str if open_punch else None
    logout_msg = f"Successfully logged out! Goodbye {user.get('name')}."
    if hours_warning:
        logout_msg += f" {hours_warning}"
    return StandardResponse(status=True, message=logout_msg, data={
        "hours_warning": hours_warning,
        "duration": duration_str_final,
    })


# ──────────────────────────────────────────────
#  GET /api/auth/attendance (Admin Only)
# ──────────────────────────────────────────────

@router.get("/attendance", response_model=StandardResponse)
async def get_attendance_logs(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Fetch real-time attendance logs for all employees (Admin only)."""
    try:
        payload = await get_current_user(credentials)
        user_id = payload.get("sub")
        if not user_id or not isinstance(user_id, str):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
        
        from app.services.auth_service import AuthService
        auth_service = AuthService(db)
        user_doc = await auth_service.get_user_by_id(user_id)
        
        if not user_doc:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="User not found."
            )

        is_admin = await _is_admin(user_doc, db)

        # Get today's date string for filtering (IST-aligned)
        now_ist = datetime.now(IST)
        today_str = now_ist.strftime("%Y-%m-%d")

        # Scope: admin sees all, employee sees only themselves
        if is_admin:
            user_filter = {}
        else:
            user_filter = {"employee_id": user_doc.get("employee_id")}

        # Fetch employees based on scope
        users = await db.users.find(user_filter, {
            "name": 1, 
            "employee_id": 1, 
            "designation": 1,
            "role": 1,
            "profession": 1,
            "hours_per_day": 1,
        }).to_list(1000)

        # Fetch today's attendance punch records
        today_punches = await db.attendance.find({
            "date": today_str,
            "type": "attendance_punch"
        }).to_list(1000)

        # Build lookup: employee_id -> list of all punches for today
        punch_map = {}
        for p in today_punches:
            eid = p.get("employee_id")
            if eid not in punch_map:
                punch_map[eid] = []
            punch_map[eid].append(p)
        
        attendance_data = []
        for u in users:
            try:
                eid = u.get("employee_id", "N/A")
                punches = punch_map.get(eid, [])

                # Determine role label
                role = u.get("role", "user")
                profession = u.get("profession", "Employee")

                if not punches:
                    # User hasn't punched in today — skip to avoid cluttering
                    continue

                # Check if employee has any completed logout today
                has_completed_logout = any(p.get("punch_out") for p in punches)
                # Check if employee has an active session (punched in but not out)
                has_active_session = any(p.get("punch_in") and not p.get("punch_out") for p in punches)

                # Add each punch as a separate row
                for punch in punches:
                    pin = punch.get("punch_in")
                    pout = punch.get("punch_out")

                    login_time_str = None
                    logout_time_str = None
                    duration_str = "-"
                    emp_status = "Absent"

                    if pin:
                        login_time_str = pin if isinstance(pin, str) else pin.isoformat()
                        emp_status = "In Office"

                    if pout:
                        logout_time_str = pout if isinstance(pout, str) else pout.isoformat()
                        emp_status = "Logged Out"

                    if punch.get("duration"):
                        duration_str = punch["duration"]
                    elif pin and not pout:
                        try:
                            pin_dt = datetime.fromisoformat(pin) if isinstance(pin, str) else pin
                            diff = now_ist - pin_dt
                            h = int(diff.total_seconds() // 3600)
                            m = int((diff.total_seconds() % 3600) // 60)
                            duration_str = f"{h}h {m}m (active)"
                        except Exception:
                            duration_str = "Active"

                    attendance_data.append({
                        "employee_id": eid,
                        "name": u.get("name", "Unknown"),
                        "designation": profession,
                        "role": role,
                        "hours_per_day": u.get("hours_per_day", 8),
                        "login_time": login_time_str,
                        "logout_time": logout_time_str,
                        "duration": duration_str,
                        "status": emp_status,
                        "login_location": punch.get("login_location"),
                        "logout_location": punch.get("logout_location"),
                        "has_completed_logout": has_completed_logout,
                        "has_active_session": has_active_session,
                    })
            except Exception as user_err:
                logger.warning(f"Skipping user {u.get('employee_id', '?')}: {user_err}")
                continue
            
        # Sort: In Office first, then Logged Out, then Absent
        status_order = {"In Office": 0, "Logged Out": 1, "Absent": 2}
        attendance_data.sort(key=lambda x: (status_order.get(x["status"], 3), x["name"]))

        return StandardResponse(status=True, message="Attendance fetched", data=attendance_data)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Attendance endpoint error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load attendance: {str(e)}")


# ──────────────────────────────────────────────
#  POST /api/auth/attendance/punch-in  (Employee marks arrival)
# ──────────────────────────────────────────────

async def _verify_punch_face_location(user_doc, face_image: Optional[str], challenge_frame: Optional[str], location: Optional[LocationData], db):
    """Helper to enforce strict Face + Location match for punches, respecting skip flags."""
    skip_face = user_doc.get("skip_face", False)
    skip_location = user_doc.get("skip_location", False)

    from app.services.auth_service import AuthService, haversine_distance, LOCATION_RADIUS_M
    auth_service = AuthService(db)

    # 1. Face Match & Liveness (skip if admin toggled skip_face)
    if not skip_face:
        if not face_image:
            raise HTTPException(status_code=400, detail="Face recognition is required for this action")
        is_verified, message, confidence = await auth_service.verify_face(
            user_id=str(user_doc["_id"]),
            face_image=face_image,
            challenge_frame=challenge_frame,
        )
        if not is_verified:
            raise HTTPException(status_code=400, detail=f"Face mismatch or spoofing detected: {message}")

    # 2. Location Match (skip if admin toggled skip_location)
    if not skip_location:
        if not location:
            raise HTTPException(status_code=400, detail="Location matching is required")
        
        reg_loc = user_doc.get("registered_location")
        login_loc = location.model_dump()
        if reg_loc and login_loc:
            dist = haversine_distance(
                reg_loc["latitude"], reg_loc["longitude"],
                login_loc["latitude"], login_loc["longitude"]
            )
            if dist > LOCATION_RADIUS_M:
                raise HTTPException(
                    status_code=400,
                    detail=f"Location Mismatch!\nRegistered: {reg_loc.get('address')}\nCurrent: {login_loc.get('address')}\nDistance: {dist:.1f}m (Max {LOCATION_RADIUS_M}m)"
                )

class PunchRequest(BaseModel):
    employee_id: str = Field(..., description="Employee ID to punch in/out")
    face_image: Optional[str] = None
    challenge_frame: Optional[str] = None
    location: Optional[LocationData] = None

@router.post("/attendance/punch-in", response_model=StandardResponse)
async def punch_in(request: PunchRequest, db=Depends(get_database)):
    """Mark employee arrival — creates a new attendance record for the day."""
    user = await db.users.find_one({"employee_id": request.employee_id})
    if not user:
        return StandardResponse(status=False, message="Employee ID not found. Please check and try again.")

    # Enforce Face + Location match
    await _verify_punch_face_location(user, request.face_image, request.challenge_frame, request.location, db)

    now_utc = datetime.now(timezone.utc)
    now_ist = now_utc.astimezone(IST)
    today_str = now_ist.strftime("%Y-%m-%d")

    # Check if already punched in today (specifically for attendance_punch type)
    existing = await db.attendance.find_one({
        "employee_id": request.employee_id,
        "date": today_str,
        "punch_out": None,
        "type": "attendance_punch"
    })
    if existing:
        return StandardResponse(
            status=False,
            message=f"Already punched in today at {existing.get('punch_in', 'unknown')}. Please punch out first."
        )

    # Geocode for address
    loc_dict = request.location.model_dump() if request.location else None
    if loc_dict:
        from app.utils.geocoding import reverse_geocode
        addr_res = await reverse_geocode(loc_dict["latitude"], loc_dict["longitude"])
        loc_dict["address"] = addr_res.get("display_name") or addr_res.get("address") or loc_dict.get("address")
        if addr_res:
            for k, v in addr_res.items():
                if k not in loc_dict:
                    loc_dict[k] = v

    record = {
        "employee_id": request.employee_id,
        "name": user.get("name", "Unknown"),
        "designation": user.get("designation", "Employee"),
        "date": today_str,
        "punch_in": now_ist.isoformat(),
        "punch_out": None,
        "duration": None,
        "status": "In Office",
        "type": "attendance_punch",
        "login_location": loc_dict,
        "created_at": now_utc
    }
    await db.attendance.insert_one(record)

    # Also update user's last_login_at
    await db.users.update_one(
        {"employee_id": request.employee_id},
        {"$set": {"last_login_at": now_utc}}
    )

    return StandardResponse(
        status=True,
        message=f"Welcome {user.get('name')}! Punched in at {now_ist.strftime('%I:%M %p')} IST.",
        data={
            "employee_id": request.employee_id,
            "name": user.get("name"),
            "punch_in": now_ist.isoformat(),
            "date": today_str
        }
    )


# ──────────────────────────────────────────────
#  POST /api/auth/attendance/verify-logout  (Pre-logout validation)
# ──────────────────────────────────────────────

@router.post("/attendance/verify-logout", response_model=StandardResponse)
async def verify_logout(request: PunchRequest, db=Depends(get_database)):
    """Verifies face and location before allowing redirection to Logout Kiosk."""
    user = await db.users.find_one({"employee_id": request.employee_id})
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    await _verify_punch_face_location(
        user_doc=user,
        face_image=request.face_image,
        challenge_frame=request.challenge_frame,
        location=request.location,
        db=db
    )
    
    return StandardResponse(status=True, message="Face and location verified successfully")

# ──────────────────────────────────────────────
#  POST /api/auth/attendance/punch-out  (Employee marks departure)
# ──────────────────────────────────────────────

@router.post("/attendance/punch-out", response_model=StandardResponse)
async def punch_out(request: PunchRequest, db=Depends(get_database)):
    """Mark employee departure — closes the open attendance record for today."""
    user = await db.users.find_one({"employee_id": request.employee_id})
    if not user:
        return StandardResponse(status=False, message="Employee ID not found.")

    now_utc = datetime.now(timezone.utc)
    now_ist = now_utc.astimezone(IST)
    today_str = now_ist.strftime("%Y-%m-%d")

    # Find open punch-in for today
    existing = await db.attendance.find_one({
        "employee_id": request.employee_id,
        "date": today_str,
        "punch_out": None,
        "type": "attendance_punch"
    })
    if not existing:
        return StandardResponse(
            status=False,
            message="No active punch-in found for today. Please punch in first."
        )

    # Calculate duration
    punch_in_str = existing.get("punch_in", "")
    try:
        punch_in_dt = datetime.fromisoformat(punch_in_str)
        # Ensure both are timezone-aware for subtraction
        if punch_in_dt.tzinfo is None:
            punch_in_dt = punch_in_dt.replace(tzinfo=IST)
        diff = now_ist - punch_in_dt
        hours = int(diff.total_seconds() // 3600)
        minutes = int((diff.total_seconds() % 3600) // 60)
        duration_str = f"{hours}h {minutes}m"
    except (ValueError, TypeError):
        duration_str = "N/A"

    # Geocode for address
    loc_dict = request.location.model_dump() if request.location else None
    if loc_dict:
        from app.utils.geocoding import reverse_geocode
        addr_res = await reverse_geocode(loc_dict["latitude"], loc_dict["longitude"])
        loc_dict["address"] = addr_res.get("display_name") or addr_res.get("address") or loc_dict.get("address")
        if addr_res:
            for k, v in addr_res.items():
                if k not in loc_dict:
                    loc_dict[k] = v

    await db.attendance.update_one(
        {"_id": existing["_id"]},
        {"$set": {
            "punch_out": now_ist.isoformat(),
            "duration": duration_str,
            "status": "Logged Out",
            "logout_location": loc_dict
        }}
    )

    # Update user's last_logout_at
    await db.users.update_one(
        {"employee_id": request.employee_id},
        {"$set": {"last_logout_at": now_utc}}
    )

    return StandardResponse(
        status=True,
        message=f"Goodbye {user.get('name')}! Punched out at {now_ist.strftime('%I:%M %p')} IST. Total: {duration_str}.",
        data={
            "employee_id": request.employee_id,
            "name": user.get("name"),
            "punch_out": now_ist.isoformat(),
            "duration": duration_str
        }
    )


# ──────────────────────────────────────────────
#  GET /api/auth/attendance/status/{employee_id}
# ──────────────────────────────────────────────

@router.get("/attendance/status/{employee_id}", response_model=StandardResponse)
async def get_attendance_status(employee_id: str, db=Depends(get_database)):
    """Check today's attendance status for an employee."""
    user = await db.users.find_one({"employee_id": employee_id})
    if not user:
        return StandardResponse(status=False, message="Employee ID not found.")

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    record = await db.attendance.find_one({
        "employee_id": employee_id,
        "date": today_start.isoformat()[:10],
        "type": "attendance_punch"
    }, sort=[("created_at", -1)])

    if not record:
        return StandardResponse(status=True, message="Not punched in today", data={
            "employee_id": employee_id,
            "name": user.get("name"),
            "designation": user.get("designation"),
            "status": "Absent",
            "punch_in": None,
            "punch_out": None,
            "duration": None
        })

    return StandardResponse(status=True, message="Status retrieved", data={
        "employee_id": employee_id,
        "name": user.get("name"),
        "designation": user.get("designation"),
        "status": "In Office" if not record.get("punch_out") else "Left Office",
        "punch_in": record.get("punch_in"),
        "punch_out": record.get("punch_out"),
        "duration": record.get("duration")
    })


# ──────────────────────────────────────────────
#  POST /api/auth/geocode
# ──────────────────────────────────────────────

class GeocodeRequest(BaseModel):
    latitude: float
    longitude: float


@router.post("/geocode")
async def geocode_location(request: GeocodeRequest):
    """Reverse-geocode a lat/lng to a human-readable address."""
    from app.utils.geocoding import reverse_geocode
    result = await reverse_geocode(request.latitude, request.longitude)
    return {"status": True, "data": result}


# ──────────────────────────────────────────────
#  GET /api/auth/attendance/report/consolidated
#  Monthly consolidated report (Admin only)
# ──────────────────────────────────────────────

@router.get("/attendance/report/consolidated", response_model=StandardResponse)
async def get_consolidated_report(
    month: Optional[int] = None,
    year: Optional[int] = None,
    department: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """
    Consolidated monthly attendance report.
    Shows: employee name, dept, total days in month, days present, days absent, total hours.
    """
    try:
        payload = await get_current_user(credentials)
        user_id = payload.get("sub")
        if not user_id or not isinstance(user_id, str):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
        auth_service = AuthService(db)
        user_doc = await auth_service.get_user_by_id(user_id)
        
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")

        is_admin = await _is_admin(user_doc, db)

        now = datetime.now(timezone.utc)
        target_month = month or now.month
        target_year = year or now.year

        import calendar
        days_in_month = calendar.monthrange(target_year, target_month)[1]

        # Build date range strings for the month
        date_prefix = f"{target_year}-{str(target_month).zfill(2)}"

        settings_doc = await db.settings.find_one({"type": "company_config"}) or {}
        office_start_time = settings_doc.get("office_start_time", "09:00")
        grace_period_mins = settings_doc.get("grace_period_mins", 30)

        # Scope: non-admin employees only see themselves
        user_filter = {}
        if not is_admin:
            user_filter["employee_id"] = user_doc.get("employee_id")
        elif department:
            user_filter["$or"] = [
                {"profession": {"$regex": department, "$options": "i"}},
                {"designation": {"$regex": department, "$options": "i"}},
                {"name": {"$regex": department, "$options": "i"}},
                {"employee_id": {"$regex": department, "$options": "i"}},
            ]

        users = await db.users.find(user_filter, {
            "name": 1, "employee_id": 1, "profession": 1, "designation": 1, "hours_per_day": 1, "weekly_off": 1
        }).to_list(1000)

        # Get global settings fallback
        settings_doc = await db.settings.find_one({"type": "company_config"})
        global_weekly_off = settings_doc.get("weekly_off", "Sunday") if settings_doc else "Sunday"

        # Fetch all attendance punches for this month
        all_punches = await db.attendance.find({
            "type": "attendance_punch",
            "date": {"$regex": f"^{date_prefix}"}
        }).to_list(10000)

        # Build lookup: employee_id -> set of dates present
        punch_by_employee = {}
        hours_by_employee = {}
        late_sums_by_employee = {}
        early_sums_by_employee = {}
        for p in all_punches:
            eid = p.get("employee_id")
            date = p.get("date")
            if not eid or not date:
                continue
            punch_by_employee.setdefault(eid, set()).add(date)
            
            # Late / Early mins
            login_t = p.get("punch_in", "")
            _, _, l_mins, e_mins = calc_late_early(login_t, office_start_time, grace_period_mins)
            
            # Use dictionaries to store totals
            if "late" not in p:
                late_sums_by_employee[eid] = late_sums_by_employee.get(eid, 0) + l_mins
                early_sums_by_employee[eid] = early_sums_by_employee.get(eid, 0) + e_mins
            
            # Accumulate hours
            dur = p.get("duration", "")
            if dur and "h" in dur:
                try:
                    parts = dur.replace("(active)", "").strip().split("h")
                    h = int(parts[0].strip())
                    m = int(parts[1].replace("m", "").strip()) if len(parts) > 1 and parts[1].strip().replace("m","").strip() else 0
                    hours_by_employee[eid] = hours_by_employee.get(eid, 0) + h + (m / 60)
                except Exception:
                    pass

        report = []
        for u in users:
            eid = u.get("employee_id", "N/A")
            present_dates = punch_by_employee.get(eid, set())
            days_present = len(present_dates)
            total_hours = round(hours_by_employee.get(eid, 0), 1)

            emp_weekly_off = u.get("weekly_off") or global_weekly_off
            
            # Count how many weekly_off days are in this month
            weekly_off_count = sum(
                1 for day in range(1, days_in_month + 1)
                if datetime(target_year, target_month, day).strftime("%A").lower() == emp_weekly_off.lower()
            )

            # Days absent = Days in month - Days present - Weekly Offs
            days_absent = max(0, days_in_month - days_present - weekly_off_count)
            target_hours_monthly = round(u.get("hours_per_day", 8) * (days_in_month - weekly_off_count), 1)
            target_hours_daily = u.get("hours_per_day", 8)

            diff = total_hours - target_hours_monthly
            overtime_hours = round(diff, 1) if diff > 0 else 0.0
            deficit_hours = round(abs(diff), 1) if diff < 0 else 0.0

            def format_m(m):
                if m == 0: return ""
                h = m // 60
                mins = m % 60
                if h > 0: return f"{h}h {mins}m"
                return f"{mins}m"

            late_total = late_sums_by_employee.get(eid, 0)
            early_total = early_sums_by_employee.get(eid, 0)

            report.append({
                "employee_id": eid,
                "name": u.get("name", "Unknown"),
                "department": u.get("profession", "Employee"),
                "designation": u.get("designation", ""),
                "days_in_month": days_in_month,
                "days_present": days_present,
                "days_absent": days_absent,
                "total_hours": total_hours,
                "target_hours": target_hours_daily,
                "overtime_hours": overtime_hours,
                "deficit_hours": deficit_hours,
                "total_late": format_m(late_total),
                "total_early": format_m(early_total),
            })

        report.sort(key=lambda x: x["name"])

        return StandardResponse(status=True, message="Consolidated report", data={
            "month": target_month,
            "year": target_year,
            "days_in_month": days_in_month,
            "total_employees": len(report),
            "department_filter": department,
            "employees": report
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Consolidated report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────────────
#  GET /api/auth/attendance/report/details
#  Detailed daily attendance report (Admin only)
# ──────────────────────────────────────────────

@router.get("/attendance/report/details", response_model=StandardResponse)
async def get_details_report(
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None,
    department: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """
    Detailed daily attendance report.
    Shows per-employee, per-day: login time, logout time, hours worked, present/absent.
    """
    try:
        payload = await get_current_user(credentials)
        user_id = payload.get("sub")
        if not user_id or not isinstance(user_id, str):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
        auth_service = AuthService(db)
        user_doc = await auth_service.get_user_by_id(user_id)
        
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")

        is_admin = await _is_admin(user_doc, db)

        now = datetime.now(timezone.utc)
        target_month = month or now.month
        target_year = year or now.year

        import calendar
        days_in_month = calendar.monthrange(target_year, target_month)[1]
        date_prefix = f"{target_year}-{str(target_month).zfill(2)}"

        # Scope: non-admin employees only see themselves
        user_filter = {}
        if not is_admin:
            user_filter["employee_id"] = user_doc.get("employee_id")
        else:
            if employee_id:
                user_filter["employee_id"] = employee_id
            if department:
                dept_filter = [
                    {"profession": {"$regex": department, "$options": "i"}},
                    {"designation": {"$regex": department, "$options": "i"}},
                    {"name": {"$regex": department, "$options": "i"}},
                    {"employee_id": {"$regex": department, "$options": "i"}},
                ]
                if "employee_id" in user_filter:
                    user_filter = {"$and": [{"employee_id": user_filter["employee_id"]}, {"$or": dept_filter}]}
                else:
                    user_filter["$or"] = dept_filter

        users = await db.users.find(user_filter, {
            "name": 1, "employee_id": 1, "profession": 1, "designation": 1, "weekly_off": 1
        }).to_list(1000)

        # Get global settings fallback
        settings_doc = await db.settings.find_one({"type": "company_config"})
        global_weekly_off = settings_doc.get("weekly_off", "Sunday") if settings_doc else "Sunday"

        # Fetch all punches for the month
        all_punches = await db.attendance.find({
            "type": "attendance_punch",
            "date": {"$regex": f"^{date_prefix}"}
        }).to_list(10000)

        # Build lookup: (employee_id, date) -> punch record
        punch_map = {}
        for p in all_punches:
            key = (p.get("employee_id"), p.get("date"))
            punch_map[key] = p

        settings_doc = await db.settings.find_one({"type": "company_config"}) or {}
        office_start_time = settings_doc.get("office_start_time", "09:00")
        grace_period_mins = settings_doc.get("grace_period_mins", 30)

        report = []
        for u in users:
            eid = u.get("employee_id", "N/A")
            daily_records = []
            
            for day in range(1, days_in_month + 1):
                date_str = f"{target_year}-{str(target_month).zfill(2)}-{str(day).zfill(2)}"
                punch = punch_map.get((eid, date_str))

                if punch:
                    # Calculate daily OT/Deficit
                    emp_hours_per_day = u.get("hours_per_day", 8)
                    daily_ot = 0.0
                    daily_def = 0.0
                    
                    dur_str = punch.get("duration", "")
                    if dur_str and "h" in dur_str:
                        try:
                            parts = dur_str.replace("(active)", "").strip().split("h")
                            h = int(parts[0].strip())
                            m = int(parts[1].replace("m", "").strip()) if len(parts) > 1 and parts[1].strip().replace("m","").strip() else 0
                            daily_worked = h + (m / 60)
                            
                            diff = daily_worked - emp_hours_per_day
                            daily_ot = round(diff, 1) if diff > 0 else 0.0
                            daily_def = round(abs(diff), 1) if diff < 0 else 0.0
                        except:
                            pass
                            
                    late_s, early_s, _, _ = calc_late_early(punch.get("punch_in", ""), office_start_time, grace_period_mins)

                    daily_records.append({
                        "date": date_str,
                        "day": datetime(target_year, target_month, day).strftime("%A"),
                        "status": "Present",
                        "login_time": punch.get("punch_in"),
                        "logout_time": punch.get("punch_out"),
                        "duration": punch.get("duration", "-"),
                        "overtime_hours": daily_ot,
                        "deficit_hours": daily_def,
                        "late_by": late_s,
                        "early_by": early_s,
                        "login_location": punch.get("login_location"),
                        "logout_location": punch.get("logout_location"),
                    })
                else:
                    day_name = datetime(target_year, target_month, day).strftime("%A")
                    emp_weekly_off = u.get("weekly_off") or global_weekly_off
                    is_off = day_name.lower() == emp_weekly_off.lower()

                    # Only mark as absent if date is in the past or today
                    date_obj = datetime(target_year, target_month, day)
                    if date_obj.date() <= now.date():
                        day_status = "Weekly Off" if is_off else "Absent"
                        daily_records.append({
                            "date": date_str,
                            "day": day_name,
                            "status": day_status,
                            "login_time": None,
                            "logout_time": None,
                            "duration": "-",
                            "overtime_hours": 0.0,
                            "deficit_hours": 0.0,
                            "late_by": "",
                            "early_by": "",
                        })
                    else:
                        daily_records.append({
                            "date": date_str,
                            "day": day_name,
                            "status": "Upcoming",
                            "login_time": None,
                            "logout_time": None,
                            "duration": "-",
                            "overtime_hours": 0.0,
                            "deficit_hours": 0.0,
                            "late_by": "",
                            "early_by": "",
                        })

            report.append({
                "employee_id": eid,
                "name": u.get("name", "Unknown"),
                "department": u.get("profession", "Employee"),
                "designation": u.get("designation", ""),
                "daily": daily_records,
            })

        report.sort(key=lambda x: x["name"])

        return StandardResponse(status=True, message="Details report", data={
            "month": target_month,
            "year": target_year,
            "days_in_month": days_in_month,
            "employees": report
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Details report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/attendance/daily-status", response_model=StandardResponse)
async def get_daily_status(
    date_str: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """
    Get daily login/logout status and locations for a specific date.
    - Admin: sees all employees' status.
    - Regular employee: sees only their own status.
    """
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")

    from app.services.auth_service import AuthService
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")

    is_admin = await _is_admin(user_doc, db)

    # 1. Parse target date
    now_ist = datetime.now(timezone.utc).astimezone(IST)
    if not date_str:
        target_date_str = now_ist.strftime("%Y-%m-%d")
    else:
        target_date_str = date_str.strip()

    try:
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    today_date = now_ist.date()
    day_name = datetime(target_date.year, target_date.month, target_date.day).strftime("%A")

    # Get global weekly off
    global_weekly_off = "Sunday"
    settings_doc = await db.settings.find_one({"type": "company_config"})
    if settings_doc:
        global_weekly_off = settings_doc.get("weekly_off", "Sunday")

    # 2. Fetch users
    if is_admin:
        users_list = await db.users.find({}, {"employee_id": 1, "name": 1, "login_sessions": 1, "weekly_off": 1, "profession": 1, "registered_location": 1}).to_list(1000)
    else:
        users_list = [user_doc]

    # 3. Fetch attendance punch records for this date
    emp_ids = [u.get("employee_id") for u in users_list if u.get("employee_id")]
    attendance_records = await db.attendance.find({
        "employee_id": {"$in": emp_ids},
        "date": target_date_str
    }).to_list(1000)

    # Map attendance records by employee_id
    att_map = {}
    for r in attendance_records:
        eid = r.get("employee_id")
        if eid:
            att_map[eid] = r

    result = []
    for u in users_list:
        eid = u.get("employee_id")
        if not eid:
            continue

        emp_weekly_off = u.get("weekly_off") or global_weekly_off
        is_weekly_off = day_name.lower() == emp_weekly_off.lower()

        punch_in = None
        punch_out = None
        login_location = None
        logout_location = None
        worked_hours = None
        status = ""

        # First check the main attendance record for today YYYY-MM-DD
        att_record = att_map.get(eid)
        if att_record:
            punch_in = att_record.get("punch_in")
            punch_out = att_record.get("punch_out")
            login_location = att_record.get("login_location")
            logout_location = att_record.get("logout_location")
            worked_hours = att_record.get("worked_hours")
            status = att_record.get("status") or "In Office"
        
        # Also check user's login_sessions for any session on this target date
        sessions = u.get("login_sessions") or []
        session_matches = []
        for s in sessions:
            login_at = s.get("login_at")
            if not login_at:
                continue
            if isinstance(login_at, datetime):
                login_at_str = login_at.isoformat()
            else:
                login_at_str = str(login_at)
            
            if login_at_str.startswith(target_date_str):
                session_matches.append(s)

        if session_matches:
            session_matches.sort(key=lambda x: x.get("login_at") or "")
            first_sess = session_matches[0]
            last_sess = session_matches[-1]

            # Merge coordinates & address for login
            login_loc_merged = None
            if first_sess.get("location"):
                login_loc_merged = {**first_sess["location"]}
                if first_sess.get("address"):
                    login_loc_merged.update(first_sess["address"])

            # Merge coordinates & address for logout
            logout_loc_merged = None
            if last_sess.get("logout_location"):
                logout_loc_merged = {**last_sess["logout_location"]}
                if last_sess.get("logout_address"):
                    logout_loc_merged.update(last_sess["logout_address"])

            # Read times
            first_login = first_sess.get("login_at")
            if isinstance(first_login, datetime):
                first_login = first_login.isoformat()
            last_logout = last_sess.get("logout_at")
            if isinstance(last_logout, datetime):
                last_logout = last_logout.isoformat()

            # Set fields if not already populated from attendance_punch
            if not punch_in:
                punch_in = first_login
                login_location = login_loc_merged
            if not punch_out:
                punch_out = last_logout
                logout_location = logout_loc_merged

            # If logged out, status is "Completed / Left Office", else "Present / In Office"
            if punch_out:
                status = "Left Office"
            else:
                status = "In Office"

            # Calculate worked hours if needed
            if punch_in and punch_out and worked_hours is None:
                try:
                    pin_dt = datetime.fromisoformat(punch_in)
                    pout_dt = datetime.fromisoformat(punch_out)
                    if pin_dt.tzinfo is None:
                        pin_dt = pin_dt.replace(tzinfo=timezone.utc)
                    if pout_dt.tzinfo is None:
                        pout_dt = pout_dt.replace(tzinfo=timezone.utc)
                    worked_hours = round((pout_dt - pin_dt).total_seconds() / 3600, 2)
                except Exception:
                    pass

        # Determine no-session status labels
        if not punch_in:
            if is_weekly_off:
                status = "Weekly Off"
            elif target_date < today_date:
                status = "Absent"
            elif target_date == today_date:
                status = "Not Logged In"
            else:
                status = "Upcoming"

        result.append({
            "employee_id": eid,
            "employee_name": u.get("name", "Unknown"),
            "profession": u.get("profession", "Employee"),
            "date": target_date_str,
            "punch_in": punch_in,
            "punch_out": punch_out,
            "login_location": login_location,
            "logout_location": logout_location,
            "worked_hours": worked_hours,
            "status": status,
            "registered_location": u.get("registered_location")
        })

    return StandardResponse(status=True, message=f"Daily status for {target_date_str} fetched", data=result)


@router.post("/admin/employee/{employee_id}/upload-document", response_model=StandardResponse)
async def admin_upload_employee_document(
    employee_id: str,
    document_type: str = Form(...),
    file: UploadFile = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Upload Government ID or Insurance document for an employee. Admin only."""
    from fastapi import HTTPException, status
    import base64
    from datetime import datetime, timezone
    from bson import ObjectId
    from app.services.auth_service import AuthService
    
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    # Read and validate file size
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File too large (limit 10MB)")

    base64_data = base64.b64encode(content).decode("utf-8")
    
    if ObjectId.is_valid(employee_id):
        query = {"_id": ObjectId(employee_id)}
    else:
        query = {"employee_id": employee_id}
        
    doc_field = "govt_id_card" if document_type == "govt_id" else "insurance_card"
    
    update_data = {
        doc_field: {
            "filename": file.filename,
            "content_type": file.content_type,
            "data": base64_data
        },
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.users.update_one(query, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
        
    return StandardResponse(status=True, message=f"{document_type.replace('_', ' ').capitalize()} uploaded successfully")


@router.get("/admin/employee/{employee_id}/download-document/{document_type}")
async def admin_download_employee_document(
    employee_id: str,
    document_type: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db=Depends(get_database)
):
    """Download Government ID or Insurance document for an employee. Admin only."""
    from fastapi import HTTPException, status, Response
    import base64
    from bson import ObjectId
    from app.services.auth_service import AuthService
    
    payload = await get_current_user(credentials)
    user_id = payload.get("sub")
    if not user_id or not isinstance(user_id, str):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    auth_service = AuthService(db)
    user_doc = await auth_service.get_user_by_id(user_id)
    if not user_doc or not await _is_admin(user_doc, db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    if ObjectId.is_valid(employee_id):
        query = {"_id": ObjectId(employee_id)}
    else:
        query = {"employee_id": employee_id}
        
    target_user = await db.users.find_one(query)
    if not target_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
        
    doc_field = "govt_id_card" if document_type == "govt_id" else "insurance_card"
    doc_obj = target_user.get(doc_field)
    
    if not doc_obj or not doc_obj.get("data"):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"No {document_type.replace('_', ' ')} uploaded for this employee")
        
    file_bytes = base64.b64decode(doc_obj["data"])
    filename = doc_obj.get("filename", f"{document_type}_document")
    content_type = doc_obj.get("content_type", "application/octet-stream")
    
    return Response(
        content=file_bytes,
        media_type=content_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


